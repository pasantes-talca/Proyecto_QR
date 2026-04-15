import os, sys, json, logging
from datetime import date, datetime

import ttkbootstrap as tb
from ttkbootstrap.constants import *
from tkinter import messagebox, StringVar, filedialog

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    psycopg2 = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    openpyxl_ok = True
except ImportError:
    openpyxl_ok = False

# ── LOGGING ───────────────────────────────────────────────────────────────────
def _log_file_path():
    base = os.path.dirname(sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))
    return os.path.join(base, "defectos.log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(_log_file_path(), encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ── CONFIG ────────────────────────────────────────────────────────────────────
def get_app_dir():
    return os.path.dirname(sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))

APP_DIR     = get_app_dir()
CONFIG_FILE = os.path.join(APP_DIR, "config.json")

DEFAULT_PG = {
    "host":"10.242.4.13",
    "port": 5432,
    "dbname": "stock",
    "user": "postgres",
    "password": "Talca2025",
    "client_encoding": "WIN1252",
    "schema": "produccion",
    "table_products": "productos",
    "table_stock": "stock",
    "table_bajas": "bajas",
    "table_sheet": "sheet",
}

MOTIVOS = [
    "Sin gas",
    "Roto",
    "Etiqueta deste\u00f1ida",
    "Envase deforme",
    "Problema Tapas",
    "Sabor",
]

def load_config():
    if not os.path.exists(CONFIG_FILE):
        return {}
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception as exc:
        log.warning("No se pudo leer config.json: %s", exc)
        return {}

def get_pg_config():
    cfg  = DEFAULT_PG.copy()
    data = load_config()
    if isinstance(data.get("pg"), dict):
        for k, v in data["pg"].items():
            if v is not None and str(v).strip():
                cfg[k] = v
    try:
        cfg["port"] = int(cfg["port"])
    except Exception:
        cfg["port"] = 5432
    return cfg

def pg_connect():
    if psycopg2 is None:
        raise RuntimeError("Falta psycopg2. Instala con: pip install psycopg2-binary")
    cfg  = get_pg_config()
    conn = psycopg2.connect(
        host=cfg["host"], port=cfg["port"],
        dbname=cfg["dbname"], user=cfg["user"], password=cfg["password"],
    )
    conn.autocommit = False
    log.info("Conectado a %s:%s/%s", cfg["host"], cfg["port"], cfg["dbname"])
    return conn

# ── DB HELPERS ────────────────────────────────────────────────────────────────
def get_all_products(conn):
    cfg    = get_pg_config()
    schema = cfg["schema"]
    tprod  = cfg["table_products"]
    with conn.cursor() as cur:
        cur.execute(f"SELECT id, descripcion FROM {schema}.{tprod} ORDER BY id ASC;")
        return cur.fetchall()

def registrar_defecto(conn, id_producto, cantidad, lote, motivo):
    cfg      = get_pg_config()
    schema   = cfg["schema"]
    tdef     = cfg["table_defectos"]
    hoy      = date.today()
    lote_val = int(lote) if lote and str(lote).strip().isdigit() else None
    with conn.cursor() as cur:
        cur.execute(
            f"""
            INSERT INTO {schema}.{tdef}
                (fecha, id_producto, cantidad, lote, motivo)
            VALUES (%s, %s, %s, %s, %s);
            """,
            (hoy, int(id_producto), int(cantidad), lote_val, motivo),
        )
    conn.commit()
    log.info("Defecto registrado: pid=%s cant=%s lote=%s motivo=%s",
             id_producto, cantidad, lote_val, motivo)

def get_reporte_defectos(conn, fecha_inicio, fecha_fin):
    cfg    = get_pg_config()
    schema = cfg["schema"]
    tdef   = cfg["table_defectos"]
    tprod  = cfg["table_products"]
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT p.descripcion, d.motivo, SUM(d.cantidad) AS total
            FROM {schema}.{tdef} d
            INNER JOIN {schema}.{tprod} p ON p.id = d.id_producto
            WHERE d.fecha BETWEEN %s AND %s
            GROUP BY p.descripcion, d.motivo
            ORDER BY p.descripcion ASC, d.motivo ASC;
        """, (fecha_inicio, fecha_fin))
        return cur.fetchall()

# ── GENERADOR EXCEL ───────────────────────────────────────────────────────────
def generar_excel_reporte(filepath, filas, fecha_inicio, fecha_fin):
    wb = Workbook()
    ws = wb.active
    ws.title = "Defectos"

    AZUL_OSC  = "1A5276"
    AZUL_FILA = "D6EAF8"
    VERDE     = "D5F5E3"
    BLANCO    = "FFFFFF"

    borde_fino  = Side(style="thin",   color="AAB7B8")
    borde_medio = Side(style="medium", color="1A5276")
    border_full = Border(left=borde_fino, right=borde_fino,
                         top=borde_fino,  bottom=borde_fino)
    border_top  = Border(left=borde_fino, right=borde_fino,
                         top=borde_medio, bottom=borde_fino)

    f_titulo   = Font(name="Arial", size=16, bold=True, color=AZUL_OSC)
    f_subtit   = Font(name="Arial", size=10, color="555555")
    f_header   = Font(name="Arial", size=10, bold=True, color=BLANCO)
    f_prod     = Font(name="Arial", size=10, bold=True, color=AZUL_OSC)
    f_dato     = Font(name="Arial", size=10)
    f_subtotal = Font(name="Arial", size=10, bold=True)
    f_total    = Font(name="Arial", size=11, bold=True, color=BLANCO)

    alin_centro = Alignment(horizontal="center", vertical="center")
    alin_izq    = Alignment(horizontal="left",   vertical="center", indent=1)
    alin_der    = Alignment(horizontal="right",  vertical="center")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Reporte de Defectos"
    ws["A1"].font = f_titulo
    ws["A1"].alignment = alin_centro

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Periodo: {fecha_inicio.strftime('%d/%m/%Y')}  al  {fecha_fin.strftime('%d/%m/%Y')}"
    ws["A2"].font = f_subtit
    ws["A2"].alignment = alin_centro

    ws.merge_cells("A3:C3")
    ws["A3"] = f"Generado el: {date.today().strftime('%d/%m/%Y')}"
    ws["A3"].font = f_subtit
    ws["A3"].alignment = alin_centro

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.append([])

    row_header = 5
    for col, h in enumerate(["Producto", "Motivo", "Cantidad"], start=1):
        cell = ws.cell(row=row_header, column=col, value=h)
        cell.font = f_header
        cell.fill = PatternFill("solid", fgColor=AZUL_OSC)
        cell.alignment = alin_centro
        cell.border = border_full
    ws.row_dimensions[row_header].height = 20

    if not filas:
        ws.merge_cells("A6:C6")
        ws["A6"] = "No se encontraron defectos en el periodo seleccionado."
        ws["A6"].font = Font(name="Arial", size=10, italic=True, color="888888")
        ws["A6"].alignment = alin_centro
    else:
        productos = {}
        for desc, motivo, total in filas:
            productos.setdefault(desc, []).append((motivo, int(total)))

        current_row   = row_header + 1
        subtotal_rows = []
        alt = False

        for desc, detalle in productos.items():
            first_data_row = current_row

            for motivo, cant in detalle:
                bg  = AZUL_FILA if alt else BLANCO
                alt = not alt

                c_prod = ws.cell(row=current_row, column=1,
                                 value=desc if motivo == detalle[0][0] else "")
                c_prod.font      = f_prod if motivo == detalle[0][0] else f_dato
                c_prod.fill      = PatternFill("solid", fgColor=bg)
                c_prod.alignment = alin_izq
                c_prod.border    = border_full

                c_mot = ws.cell(row=current_row, column=2, value=motivo)
                c_mot.font      = f_dato
                c_mot.fill      = PatternFill("solid", fgColor=bg)
                c_mot.alignment = alin_izq
                c_mot.border    = border_full

                c_cant = ws.cell(row=current_row, column=3, value=cant)
                c_cant.font          = f_dato
                c_cant.fill          = PatternFill("solid", fgColor=bg)
                c_cant.alignment     = alin_der
                c_cant.border        = border_full
                c_cant.number_format = "#,##0"

                ws.row_dimensions[current_row].height = 18
                current_row += 1

            sub_row = current_row
            subtotal_rows.append(sub_row)

            c1 = ws.cell(row=sub_row, column=1, value=f"Subtotal \u2014 {desc}")
            c1.font = f_subtotal
            c1.fill = PatternFill("solid", fgColor=VERDE)
            c1.alignment = alin_izq
            c1.border = border_top

            c2 = ws.cell(row=sub_row, column=2, value="")
            c2.fill   = PatternFill("solid", fgColor=VERDE)
            c2.border = border_top

            c3 = ws.cell(row=sub_row, column=3,
                         value=f"=SUM(C{first_data_row}:C{sub_row - 1})")
            c3.font          = f_subtotal
            c3.fill          = PatternFill("solid", fgColor=VERDE)
            c3.alignment     = alin_der
            c3.border        = border_top
            c3.number_format = "#,##0"

            ws.row_dimensions[sub_row].height = 18
            current_row += 2  # fila vacía entre productos

        tot_row = current_row
        c1 = ws.cell(row=tot_row, column=1, value="TOTAL GENERAL DE DEFECTOS")
        c1.font = f_total
        c1.fill = PatternFill("solid", fgColor=AZUL_OSC)
        c1.alignment = alin_izq
        c1.border = border_full

        c2 = ws.cell(row=tot_row, column=2, value="")
        c2.fill   = PatternFill("solid", fgColor=AZUL_OSC)
        c2.border = border_full

        c3 = ws.cell(row=tot_row, column=3,
                     value="=SUM(" + ",".join(f"C{r}" for r in subtotal_rows) + ")")
        c3.font          = f_total
        c3.fill          = PatternFill("solid", fgColor=AZUL_OSC)
        c3.alignment     = alin_der
        c3.border        = border_full
        c3.number_format = "#,##0"

        ws.row_dimensions[tot_row].height = 22

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A6"

    wb.save(filepath)
    log.info("Excel generado: %s", filepath)


# ── VENTANA REPORTE ───────────────────────────────────────────────────────────
class ReporteWindow(tb.Toplevel):
    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.title("Generar Reporte Excel")
        self.geometry("480x310")
        self.resizable(False, False)
        self._build()

    def _build(self):
        tb.Label(self, text="Reporte de Defectos por Periodo",
                 font=("Segoe UI", 14, "bold")).pack(pady=(20, 10))

        frame = tb.LabelFrame(self, text="Seleccionar periodo")
        frame.pack(padx=30, pady=5, fill="x", ipadx=8, ipady=8)

        hoy = date.today().strftime("%d/%m/%Y")

        # Fecha inicio
        tb.Label(frame, text="Fecha inicio:", font=("Segoe UI", 11)).grid(
            row=0, column=0, sticky="e", padx=10, pady=10)
        self.ini_var = StringVar(value=hoy)
        tb.Entry(frame, textvariable=self.ini_var, width=14,
                 font=("Segoe UI", 11)).grid(row=0, column=1, sticky="w", padx=10, pady=10)
        tb.Label(frame, text="(dd/mm/aaaa)", font=("Segoe UI", 9, "italic"),
                 foreground="gray").grid(row=0, column=2, sticky="w")

        # Fecha fin
        tb.Label(frame, text="Fecha fin:", font=("Segoe UI", 11)).grid(
            row=1, column=0, sticky="e", padx=10, pady=10)
        self.fin_var = StringVar(value=hoy)
        tb.Entry(frame, textvariable=self.fin_var, width=14,
                 font=("Segoe UI", 11)).grid(row=1, column=1, sticky="w", padx=10, pady=10)
        tb.Label(frame, text="(dd/mm/aaaa)", font=("Segoe UI", 9, "italic"),
                 foreground="gray").grid(row=1, column=2, sticky="w")

        self.status_var = StringVar(value="")
        tb.Label(self, textvariable=self.status_var, font=("Segoe UI", 10),
                 foreground="#555", wraplength=420).pack(pady=6, padx=20)

        self.btn = tb.Button(self, text="  GENERAR EXCEL  ",
                             bootstyle=SUCCESS, width=22,
                             command=self._on_generar)
        self.btn.pack(pady=8)

    def _on_generar(self):
        if not openpyxl_ok:
            messagebox.showerror("Falta libreria",
                                 "Instala openpyxl:\n  pip install openpyxl",
                                 parent=self)
            return
        try:
            fecha_inicio = datetime.strptime(self.ini_var.get().strip(), "%d/%m/%Y").date()
            fecha_fin    = datetime.strptime(self.fin_var.get().strip(), "%d/%m/%Y").date()
        except ValueError:
            self.status_var.set("\u26a0 Formato incorrecto. Usar dd/mm/aaaa (ej: 01/03/2026)")
            return

        if fecha_inicio > fecha_fin:
            self.status_var.set("\u26a0 La fecha inicio debe ser menor o igual a la fecha fin.")
            return

        nombre = f"reporte_defectos_{fecha_inicio}_{fecha_fin}.xlsx"
        filepath = filedialog.asksaveasfilename(
            parent=self, title="Guardar reporte Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=nombre,
        )
        if not filepath:
            return

        self.btn.config(state="disabled", text="Generando\u2026")
        self.status_var.set("Consultando base de datos\u2026")
        self.update_idletasks()

        try:
            filas = get_reporte_defectos(self.conn, fecha_inicio, fecha_fin)
            generar_excel_reporte(filepath, filas, fecha_inicio, fecha_fin)
            self.status_var.set("\u2714 Excel guardado correctamente.")
            messagebox.showinfo("Listo", f"Reporte generado:\n{filepath}", parent=self)
        except Exception as exc:
            log.exception("Error al generar Excel")
            self.status_var.set(f"\u2718 Error: {exc}")
        finally:
            self.btn.config(state="normal", text="  GENERAR EXCEL  ")


# ── UI PRINCIPAL ──────────────────────────────────────────────────────────────
class DefectosApp:
    def __init__(self, conn):
        self.conn = conn
        self.root = tb.Window(themename="minty")
        self.root.title("Registro de Defectos")
        self.root.geometry("700x600")
        self.root.resizable(True, True)
        self._build_ui()
        self.root.mainloop()

    def _build_ui(self):
        root = self.root

        tb.Label(root, text="Registro de Defectos",
                 font=("Segoe UI", 22, "bold")).pack(pady=(24, 12))

        frame = tb.Labelframe(root, text="Datos del defecto")
        frame.pack(pady=6, padx=50, fill="x", ipadx=10, ipady=10)
        frame.columnconfigure(1, weight=1)

        # Producto
        tb.Label(frame, text="Producto:", font=("Segoe UI", 11)).grid(
            row=0, column=0, sticky="e", padx=10, pady=10)
        prods = get_all_products(self.conn)
        self.prod_options = [f"{pid} \u2014 {desc}" for pid, desc in prods]
        self.prod_var = StringVar()
        self.prod_combo = tb.Combobox(frame, textvariable=self.prod_var,
                                      values=self.prod_options,
                                      state="readonly", width=55)
        self.prod_combo.grid(row=0, column=1, columnspan=2,
                             sticky="w", padx=10, pady=10)
        if self.prod_options:
            self.prod_combo.current(0)

        # Cantidad
        tb.Label(frame, text="Cantidad de botellas:", font=("Segoe UI", 11)).grid(
            row=1, column=0, sticky="e", padx=10, pady=10)
        vcmd = (root.register(lambda v: v == "" or v.isdigit()), "%P")
        self.cant_var = StringVar(value="1")
        tb.Entry(frame, textvariable=self.cant_var, width=14,
                 validate="key", validatecommand=vcmd,
                 font=("Segoe UI", 11)).grid(row=1, column=1, sticky="w",
                                             padx=10, pady=10)

        # Lote
        tb.Label(frame, text="N\u00ba de lote (opcional):", font=("Segoe UI", 11)).grid(
            row=2, column=0, sticky="e", padx=10, pady=10)
        lote_vcmd = (root.register(lambda v: v == "" or v.isdigit()), "%P")
        self.lote_var = StringVar()
        tb.Entry(frame, textvariable=self.lote_var, width=20,
                 validate="key", validatecommand=lote_vcmd,
                 font=("Segoe UI", 11)).grid(row=2, column=1, sticky="w",
                                             padx=10, pady=10)
        tb.Label(frame, text="(dejar vac\u00edo si no aplica)",
                 font=("Segoe UI", 9, "italic"),
                 foreground="gray").grid(row=2, column=2, sticky="w")

        # Motivo
        tb.Label(frame, text="Motivo del defecto:", font=("Segoe UI", 11)).grid(
            row=3, column=0, sticky="e", padx=10, pady=10)
        self.motivo_var = StringVar()
        self.motivo_combo = tb.Combobox(frame, textvariable=self.motivo_var,
                                        values=MOTIVOS, state="readonly",
                                        width=35, font=("Segoe UI", 11))
        self.motivo_combo.grid(row=3, column=1, sticky="w", padx=10, pady=10)
        self.motivo_combo.current(0)

        # Estado
        self.status_var = StringVar(value="Complete el formulario y presione REGISTRAR.")
        tb.Label(root, textvariable=self.status_var, font=("Segoe UI", 11),
                 wraplength=620, justify="left",
                 foreground="#555").pack(pady=14, padx=50)

        # Botones
        btn_frame = tb.Frame(root)
        btn_frame.pack(pady=10)

        self.btn = tb.Button(btn_frame, text="  REGISTRAR DEFECTO  ",
                             bootstyle=DANGER, width=26,
                             command=self._on_registrar)
        self.btn.grid(row=0, column=0, padx=12)

        tb.Button(btn_frame, text="  GENERAR REPORTE EXCEL  ",
                  bootstyle=SUCCESS, width=26,
                  command=self._on_reporte).grid(row=0, column=1, padx=12)

    def _on_registrar(self):
        try:
            pstr = self.prod_var.get()
            if not pstr:
                raise ValueError("Seleccione un producto.")
            id_producto = int(pstr.split(" \u2014 ")[0])
            cant_str = self.cant_var.get().strip()
            if not cant_str or not cant_str.isdigit():
                raise ValueError("La cantidad debe ser un n\u00famero entero positivo.")
            cantidad = int(cant_str)
            if cantidad <= 0:
                raise ValueError("La cantidad debe ser mayor a cero.")
            lote   = self.lote_var.get().strip()
            motivo = self.motivo_var.get().strip()
            if not motivo:
                raise ValueError("Seleccione un motivo.")
        except ValueError as exc:
            self.status_var.set(f"\u26a0 {exc}")
            return

        desc     = pstr.split(" \u2014 ", 1)[1] if " \u2014 " in pstr else pstr
        lote_txt = lote if lote else "Sin especificar"
        if not messagebox.askyesno(
            "Confirmar registro",
            f"\u00bfRegistrar el siguiente defecto?\n\n"
            f"  Producto : {desc}\n"
            f"  Cantidad : {cantidad} botella(s)\n"
            f"  Lote     : {lote_txt}\n"
            f"  Motivo   : {motivo}",
            parent=self.root,
        ):
            return

        self.btn.config(state="disabled", text="Guardando\u2026")
        self.root.update_idletasks()
        try:
            registrar_defecto(self.conn, id_producto, cantidad, lote, motivo)
            self.status_var.set(
                f"\u2714 Defecto registrado. "
                f"Producto: {desc}  |  Cantidad: {cantidad}  |  "
                f"Lote: {lote_txt}  |  Motivo: {motivo}"
            )
            self.cant_var.set("1")
            self.lote_var.set("")
            self.motivo_combo.current(0)
        except Exception as exc:
            log.exception("Error al registrar defecto")
            self.status_var.set(f"\u2718 Error: {exc}")
        finally:
            self.btn.config(state="normal", text="  REGISTRAR DEFECTO  ")

    def _on_reporte(self):
        ReporteWindow(self.root, self.conn)


# ── ENTRYPOINT ────────────────────────────────────────────────────────────────
def main():
    try:
        conn = pg_connect()
    except Exception as exc:
        messagebox.showerror("Error PostgreSQL", f"No se pudo conectar:\n{exc}")
        log.critical("Fallo de conexi\u00f3n: %s", exc)
        return
    DefectosApp(conn)

if __name__ == "__main__":
    main()